"""Bulk-add for sites — CSV and XLSX.

Lets the operator import many sites at once: one row per site with a
login URL, credentials, and an optional template id.

This module is pure logic — no Flask, no module globals — so it stays
unit-testable. app.py wires the endpoints on top:
  GET  /api/sites/csv_template   -> build_csv_template()
  GET  /api/sites/xlsx_template  -> build_xlsx_template()
  POST /api/sites/bulk_csv       -> parse_import() (accepts CSV or XLSX)

Template formats:
  - CSV  : data rows FIRST (header + example, ready to edit), with the
           instructions and the available-template list as # comment
           lines BELOW the data, so a spreadsheet app shows the
           fillable part on top instead of a wall of comments.
  - XLSX : a formatted workbook — an "Instructions" sheet and a "Sites"
           sheet with a styled header, an example row, sensible column
           widths, and a dropdown on the `template` column listing
           every valid template id.
"""
from __future__ import annotations

import csv
import io
from urllib.parse import urlparse

# The columns the operator fills in, in display order.
COLUMNS = ["name", "login_url", "username", "password", "template",
           "login_template"]

# Per-column human help — shared by both the CSV comments and the XLSX
# Instructions sheet so the two formats never drift.
COLUMN_HELP = [
    ("name", "Display name. Optional — derived from the URL host if "
             "left blank."),
    ("login_url", "The site's login page URL (or its main URL). "
                  "Required."),
    ("username", "Login username or email. Leave blank for "
                 "cookie-only sites."),
    ("password", "Login password. Leave blank for cookie-only sites."),
    ("template", "Optional. A DOWNLOAD template id — how to grab the "
                 "video. Blank = teach the site manually later."),
    ("login_template", "Optional. A LOGIN template id — how to log "
                       "in. Blank = teach the login manually later."),
]

_EXAMPLE_ROW = {
    "name": "Example Site",
    "login_url": "https://example.com/login",
    "username": "user@example.com",
    "password": "changeme",
    "template": "video_js",
    "login_template": "",
}


def _template_pairs(templates):
    """Return a clean [(id, name), ...] list from templates.list_*()."""
    out = []
    for t in templates or []:
        tid = str(t.get("id", "")).strip()
        if not tid:
            continue
        out.append((tid, str(t.get("name", "")).strip()))
    return out


# ─────────────────────────── CSV template ───────────────────────────

def build_csv_template(templates, login_templates=None) -> str:
    """CSV text for the downloadable import template.

    Data first: the header row and one example row sit at the TOP so a
    spreadsheet app opens straight onto the fillable area. Instructions
    and the available-template references follow as # comment lines,
    which the parser ignores on import.
    """
    lines = []
    a = lines.append
    # --- fillable data, on top ---
    a(",".join(COLUMNS))
    a(",".join(_EXAMPLE_ROW[c] for c in COLUMNS))
    # --- everything below is ignored on import ---
    a("#")
    a("# ^^^ EDIT ABOVE — one row per site. Delete the example row or "
      "overwrite it.")
    a("# Lines starting with # are ignored. Save as CSV and import from "
      "Bulk Import.")
    a("#")
    a("# COLUMNS")
    for col, help_text in COLUMN_HELP:
        a(f"#   {col:<15}{help_text}")
    a("#")
    a("# " + "-" * 64)
    a("# DOWNLOAD TEMPLATES  -  put the id in the `template` column")
    a("# " + "-" * 64)
    for tid, name in _template_pairs(templates):
        a(f"#   {tid:<32}{name}")
    login_pairs = _template_pairs(login_templates)
    if login_pairs:
        a("#")
        a("# " + "-" * 64)
        a("# LOGIN TEMPLATES  -  put the id in the `login_template` "
          "column")
        a("# " + "-" * 64)
        for tid, name in login_pairs:
            a(f"#   {tid:<32}{name}")
    return "\r\n".join(lines) + "\r\n"


# ────────────────────────── XLSX template ───────────────────────────

def build_xlsx_template(templates, login_templates=None) -> bytes:
    """Return .xlsx bytes for the downloadable import template.

    Two sheets: 'Sites' (fillable, styled header, example row, dropdowns
    on the template and login_template columns) and 'Instructions'
    (readable guidance + the template-id references). No formulas
    besides the auto-name preview — nothing to recalculate.

    `login_templates` is the list from
    login_templates_data.list_login_templates(); when omitted, the
    login_template column simply has no dropdown.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    pairs = _template_pairs(templates)
    login_pairs = _template_pairs(login_templates)
    wb = Workbook()

    # ---- Sheet 1: Sites (the fillable form) ----
    ws = wb.active
    ws.title = "Sites"
    header_fill = PatternFill("solid", start_color="1F2A44")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    example_font = Font(name="Arial", italic=True, color="888888", size=11)
    body_font = Font(name="Arial", size=11)
    name_font = Font(name="Arial", size=11, color="555555")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    # The `name` column shows a best-effort preview pulled from
    # `login_url`: the host between :// and the first / or ?, minus a
    # leading www. This is a single-level formula on purpose — deeply
    # nested IFs to strip every subdomain prefix overflow the formula
    # engine. The authoritative cleanup happens server-side on import
    # (derive_name() strips www./m./login./members./secure.), so a
    # blank or subdomain-y preview still imports as a clean name.
    # Leave the name cell alone, or type a name to override it.
    def _name_formula(r):
        u = f"B{r}"
        host = (f'MID({u},FIND("://",{u})+3,'
                f'MIN(IFERROR(FIND("/",{u},FIND("://",{u})+3),LEN({u})+1),'
                f'IFERROR(FIND("?",{u},FIND("://",{u})+3),LEN({u})+1))'
                f'-FIND("://",{u})-3)')
        nowww = (f'IF(LEFT(LOWER({host}),4)="www.",'
                 f'MID({host},5,LEN({host})),{host})')
        return (f'=IF(B{r}="","",PROPER(LEFT({nowww},'
                f'IFERROR(FIND(".",{nowww})-1,LEN({nowww})))))')

    # Example row — italic/grey so it's obviously a placeholder.
    ws.append([_name_formula(2)] + [_EXAMPLE_ROW[c]
                                    for c in COLUMNS[1:]])
    for cell in ws[2]:
        cell.font = example_font
        cell.border = border

    # Pre-format the next ~200 rows; the name cell carries the
    # auto-name formula so it fills as soon as a URL is typed.
    for r in range(3, 203):
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = border
        nc = ws.cell(row=r, column=1)
        nc.value = _name_formula(r)
        nc.font = name_font

    widths = {"name": 26, "login_url": 42, "username": 28,
              "password": 20, "template": 24, "login_template": 24}
    for idx, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx)
                             .column_letter].width = widths[col]

    ws.freeze_panes = "A2"  # keep the header visible while scrolling

    # Dropdown on the template column, rows 2..202. The template ids
    # are written to a hidden 'TemplateIDs' sheet and the validation
    # points at that range — this has no length limit, unlike an
    # inline list (Excel caps inline lists at 255 chars, which 90+
    # template ids blow past).
    if pairs:
        ids_sheet = wb.create_sheet("TemplateIDs")
        for i, (tid, _) in enumerate(pairs, 1):
            ids_sheet.cell(row=i, column=1, value=tid)
        ids_sheet.sheet_state = "hidden"
        last = len(pairs)
        dv = DataValidation(
            type="list",
            formula1=f"TemplateIDs!$A$1:$A${last}",
            allow_blank=True)
        dv.error = "Pick a template id from the list, or leave blank."
        dv.errorTitle = "Unknown template"
        dv.prompt = "Optional — choose a template id, or leave blank."
        dv.promptTitle = "Template"
        ws.add_data_validation(dv)
        tmpl_col = chr(ord("A") + COLUMNS.index("template"))
        dv.add(f"{tmpl_col}2:{tmpl_col}202")

    # Dropdown on the login_template column, same hidden-sheet approach.
    if login_pairs:
        lids_sheet = wb.create_sheet("LoginTemplateIDs")
        for i, (tid, _) in enumerate(login_pairs, 1):
            lids_sheet.cell(row=i, column=1, value=tid)
        lids_sheet.sheet_state = "hidden"
        llast = len(login_pairs)
        ldv = DataValidation(
            type="list",
            formula1=f"LoginTemplateIDs!$A$1:$A${llast}",
            allow_blank=True)
        ldv.error = "Pick a login template id, or leave blank."
        ldv.errorTitle = "Unknown login template"
        ldv.prompt = "Optional — choose a login template, or leave blank."
        ldv.promptTitle = "Login template"
        ws.add_data_validation(ldv)
        ltmpl_col = chr(ord("A") + COLUMNS.index("login_template"))
        ldv.add(f"{ltmpl_col}2:{ltmpl_col}202")

    # ---- Sheet 2: Instructions ----
    info = wb.create_sheet("Instructions")
    title_font = Font(name="Arial", bold=True, size=14)
    h_font = Font(name="Arial", bold=True, size=11)
    body = Font(name="Arial", size=11)
    mono_id = Font(name="Consolas", size=10)

    info["A1"] = "BulkDownloader — bulk site import"
    info["A1"].font = title_font
    rows = [
        "",
        "Fill in the 'Sites' sheet — one row per site — then save the "
        "file and import it from the Bulk Import button.",
        "You can import this .xlsx directly, or save it as .csv; both "
        "work.",
        "The example row is a placeholder — overwrite it or delete it.",
        "",
        "COLUMNS",
    ]
    for col, help_text in COLUMN_HELP:
        rows.append(f"    {col} — {help_text}")
    rows += ["", "AVAILABLE TEMPLATES",
             "    Download templates — put the id in the 'template' "
             "column.", ""]
    r = 2
    for line in rows:
        info.cell(row=r, column=1, value=line).font = (
            h_font if line in ("COLUMNS", "AVAILABLE TEMPLATES") else body)
        r += 1
    # Download-template reference table: id | name
    info.cell(row=r, column=1, value="id").font = h_font
    info.cell(row=r, column=2, value="name").font = h_font
    r += 1
    for tid, name in pairs:
        info.cell(row=r, column=1, value=tid).font = mono_id
        info.cell(row=r, column=2, value=name).font = body
        r += 1
    # Login-template reference table.
    if login_pairs:
        r += 1
        info.cell(row=r, column=1,
                  value="AVAILABLE LOGIN TEMPLATES").font = h_font
        r += 1
        info.cell(row=r, column=1,
                  value="    Put the id in the 'login_template' "
                        "column.").font = body
        r += 2
        info.cell(row=r, column=1, value="id").font = h_font
        info.cell(row=r, column=2, value="name").font = h_font
        r += 1
        for tid, name in login_pairs:
            info.cell(row=r, column=1, value=tid).font = mono_id
            info.cell(row=r, column=2, value=name).font = body
            r += 1
    info.column_dimensions["A"].width = 40
    info.column_dimensions["B"].width = 48

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ───────────────────────────── parsing ──────────────────────────────

def _rows_from_records(records):
    """Shared row-normalisation for CSV and XLSX. `records` is a list of
    (line_no, header_lower_list, value_list). Returns (rows, errors)."""
    rows: list[dict] = []
    errors: list[tuple[int, str]] = []
    if not records:
        return rows, [(0, "no data rows found")]
    _, header, _ = records[0]
    if "login_url" not in header:
        return rows, [(records[0][0],
                       "header row must include a 'login_url' column")]
    for lineno, _, vals in records[1:]:
        if not any(str(v).strip() for v in vals):
            continue
        row = {c: "" for c in COLUMNS}
        for idx, col in enumerate(header):
            if col in COLUMNS and idx < len(vals):
                row[col] = str(vals[idx]).strip()
        row["line"] = lineno
        if not row["login_url"]:
            errors.append((lineno, "missing login_url"))
            continue
        rows.append(row)
    return rows, errors


def parse_csv(text: str):
    """Parse import CSV text. Returns (rows, errors).

    Comment (#) and blank lines are skipped. The first surviving line
    must be the header and must contain a 'login_url' column. The
    header may appear anywhere — data-first or comments-first templates
    both parse.
    """
    if not text or not text.strip():
        return [], [(0, "empty file")]
    records = []
    for i, raw in enumerate(text.splitlines(), 1):
        s = raw.strip()
        if not s or s.startswith("#"):
            continue
        vals = next(csv.reader([raw]), [])
        if not records:
            header = [h.strip().lower() for h in vals]
            records.append((i, header, vals))
        else:
            records.append((i, records[0][1], vals))
    return _rows_from_records(records)


def parse_xlsx(data: bytes):
    """Parse an uploaded .xlsx import file. Returns (rows, errors).

    Reads the 'Sites' sheet if present, else the first sheet. The first
    non-empty row is the header and must contain 'login_url'.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover - openpyxl is a hard dep
        return [], [(0, f"cannot read .xlsx: {e}")]
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        return [], [(0, f"not a valid .xlsx file: {e}")]
    ws = wb["Sites"] if "Sites" in wb.sheetnames else wb[wb.sheetnames[0]]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = ["" if v is None else v for v in row]
        if not any(str(v).strip() for v in vals):
            continue
        if not records:
            header = [str(h).strip().lower() for h in vals]
            records.append((i, header, vals))
        else:
            records.append((i, records[0][1], vals))
    return _rows_from_records(records)


def parse_import(*, text: str = None, xlsx: bytes = None,
                  filename: str = ""):
    """Dispatch to the right parser. Prefers an explicit format; falls
    back to sniffing the filename extension. Returns (rows, errors)."""
    if xlsx is not None and (filename.lower().endswith(".xlsx")
                             or text is None):
        return parse_xlsx(xlsx)
    return parse_csv(text or "")


def resolve_template(value, templates):
    """Resolve an import `template` cell to a template id.

    Accepts an exact id or a template name (case-insensitive). A blank
    value is valid and means 'no template'. Returns (template_id, error)
    -- template_id is None when blank, error is None on success.
    """
    v = (value or "").strip()
    if not v:
        return None, None
    by_id, by_name = {}, {}
    for t in templates or []:
        tid = t.get("id")
        if not tid:
            continue
        by_id[str(tid).lower()] = tid
        by_name[str(t.get("name", "")).lower()] = tid
    low = v.lower()
    if low in by_id:
        return by_id[low], None
    if low in by_name:
        return by_name[low], None
    return None, f"unknown template: {value!r}"


def derive_name(url: str) -> str:
    """Best-effort display name from a URL host, for rows with no name."""
    try:
        host = urlparse(url if "://" in url else "http://" + url).hostname or ""
    except Exception:
        host = ""
    host = host.lower()
    for prefix in ("www.", "m.", "login.", "members.", "secure.",
                   "auth.", "portal.", "account."):
        if host.startswith(prefix):
            host = host[len(prefix):]
    base = host.split(".")[0] if host else ""
    return base.replace("-", " ").title() or "Imported Site"


def clean_name(raw_name: str, login_url: str) -> str:
    """Resolve the display name for an imported row.

    Uses the provided name unless it is blank OR is itself a URL — a
    common paste mistake (operator drops the URL into the name column).
    In either case the name is derived from the login URL host.
    """
    n = (raw_name or "").strip()
    if not n or "://" in n or n.lower().startswith("www."):
        return derive_name(login_url)
    return n

