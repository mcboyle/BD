"""Tests for the CSV bulk site-import module (bulk_downloader.csv_bulk).

Pure-logic tests: no app boot, no Flask, fully deterministic.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from bulk_downloader import csv_bulk

_TPLS = [
    {"id": "video_js", "name": "Video.js player"},
    {"id": "jw_player", "name": "JW Player"},
    {"id": "kvs_engine", "name": "KVS (Kernel Video Sharing) engine"},
]


def test_build_template_lists_every_id():
    out = csv_bulk.build_csv_template(_TPLS)
    for t in _TPLS:
        assert t["id"] in out, f"{t['id']} missing from template"
    # v3.62.2: data-first layout — the header row is line 1 (was a
    # wall of # comments first; that looked bad in spreadsheet apps).
    assert out.startswith("name,login_url,")  # header leads the file
    assert "login_url" in out


def test_build_template_roundtrips_through_parser():
    # The generated template must itself import cleanly: comments
    # skipped, one example data row recovered.
    out = csv_bulk.build_csv_template(_TPLS)
    rows, errors = csv_bulk.parse_csv(out)
    assert not errors
    assert len(rows) == 1
    assert rows[0]["login_url"] == "https://example.com/login"
    assert rows[0]["template"] == "video_js"


def test_parse_skips_comments_and_blank_lines():
    txt = ("# a comment\n\n"
           "name,login_url,username,password,template\n"
           "Site A,https://a.com,u,p,video_js\n"
           "# mid comment\n"
           "Site B,https://b.com,,,\n")
    rows, errors = csv_bulk.parse_csv(txt)
    assert not errors
    assert len(rows) == 2
    assert rows[0]["name"] == "Site A"
    assert rows[1]["login_url"] == "https://b.com"
    assert rows[1]["template"] == ""


def test_parse_reports_missing_login_url_per_row():
    txt = "name,login_url\nGood,https://ok.com\nBad,\n"
    rows, errors = csv_bulk.parse_csv(txt)
    assert len(rows) == 1 and rows[0]["name"] == "Good"
    assert errors and "login_url" in errors[0][1]


def test_parse_requires_a_header():
    rows, errors = csv_bulk.parse_csv("# only comments here\n")
    assert not rows and errors


def test_parse_rejects_header_without_login_url():
    rows, errors = csv_bulk.parse_csv("name,username\nx,y\n")
    assert not rows and errors


def test_parse_empty_file():
    rows, errors = csv_bulk.parse_csv("")
    assert not rows and errors


def test_parse_handles_quoted_commas():
    txt = ('name,login_url,username,password,template\n'
           '"Big, Site",https://x.com,u,"p,word",jw_player\n')
    rows, errors = csv_bulk.parse_csv(txt)
    assert not errors
    assert rows[0]["name"] == "Big, Site"
    assert rows[0]["password"] == "p,word"


def test_parse_keeps_source_line_numbers():
    txt = "# c\nname,login_url\nA,https://a.com\n"
    rows, _ = csv_bulk.parse_csv(txt)
    assert rows[0]["line"] == 3


def test_resolve_template_by_id():
    tid, err = csv_bulk.resolve_template("video_js", _TPLS)
    assert tid == "video_js" and err is None


def test_resolve_template_by_name_case_insensitive():
    tid, err = csv_bulk.resolve_template("jw player", _TPLS)
    assert tid == "jw_player" and err is None


def test_resolve_template_blank_is_valid():
    tid, err = csv_bulk.resolve_template("", _TPLS)
    assert tid is None and err is None


def test_resolve_template_unknown_errors():
    tid, err = csv_bulk.resolve_template("not_a_template", _TPLS)
    assert tid is None and err


def test_derive_name_from_url():
    assert csv_bulk.derive_name("https://www.example.com/login") == "Example"
    assert csv_bulk.derive_name("members.foo-bar.net") == "Foo Bar"
    assert csv_bulk.derive_name("") == "Imported Site"


# ── v3.62.2: XLSX template + XLSX import ────────────────────────────

def test_build_xlsx_template_is_valid_workbook():
    data = csv_bulk.build_xlsx_template(_TPLS)
    assert isinstance(data, bytes) and data[:2] == b"PK"  # zip magic
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    assert "Sites" in wb.sheetnames
    assert "Instructions" in wb.sheetnames
    ws = wb["Sites"]
    header = [c.value for c in ws[1]]
    assert header == csv_bulk.COLUMNS


def test_xlsx_template_roundtrips_through_parser():
    data = csv_bulk.build_xlsx_template(_TPLS)
    rows, errors = csv_bulk.parse_xlsx(data)
    assert not errors
    assert len(rows) == 1                       # the example row
    assert rows[0]["login_url"]                 # required field present


def test_parse_xlsx_rejects_non_xlsx():
    rows, errors = csv_bulk.parse_xlsx(b"this is not a spreadsheet")
    assert not rows and errors


def test_parse_import_dispatches_by_format():
    csv_text = csv_bulk.build_csv_template(_TPLS)
    xlsx_data = csv_bulk.build_xlsx_template(_TPLS)
    r_csv, _ = csv_bulk.parse_import(text=csv_text, filename="x.csv")
    r_xlsx, _ = csv_bulk.parse_import(xlsx=xlsx_data, filename="x.xlsx")
    assert len(r_csv) == 1 and len(r_xlsx) == 1


def test_clean_name_handles_url_as_name():
    # v3.62.2: a row where the operator pasted the URL into the name
    # column must still get a clean derived name, not a URL as a name.
    assert csv_bulk.clean_name(
        "https://auth.reptyle.com/oauth/login",
        "https://auth.reptyle.com/oauth/login") == "Reptyle"
    assert csv_bulk.clean_name(
        "", "https://members.naughtyamerica.com/login") == "Naughtyamerica"
    assert csv_bulk.clean_name("My Site", "https://x.com") == "My Site"


# ── v3.62.2: login_template column ──────────────────────────────────

def test_columns_include_login_template():
    assert "login_template" in csv_bulk.COLUMNS
    assert "template" in csv_bulk.COLUMNS


def test_csv_template_lists_login_templates():
    login_tpls = [{"id": "login_foo", "name": "Foo (login)"}]
    out = csv_bulk.build_csv_template(_TPLS, login_tpls)
    assert "login_template" in out          # header column
    assert "login_foo" in out               # in the reference list
    assert "LOGIN TEMPLATES" in out


def test_xlsx_template_with_login_templates_roundtrips():
    login_tpls = [{"id": "login_foo", "name": "Foo (login)"}]
    data = csv_bulk.build_xlsx_template(_TPLS, login_tpls)
    rows, errors = csv_bulk.parse_xlsx(data)
    assert not errors and len(rows) == 1
    # the login_template column exists on the parsed row
    assert "login_template" in rows[0]


def test_resolve_template_works_for_login_templates():
    login_tpls = [{"id": "login_bar", "name": "Bar (login)"}]
    tid, err = csv_bulk.resolve_template("login_bar", login_tpls)
    assert tid == "login_bar" and err is None
    tid, err = csv_bulk.resolve_template("", login_tpls)
    assert tid is None and err is None
