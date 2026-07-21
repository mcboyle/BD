"""Pin test for tools/tasktracker_gen.py — the single-source tracker generator.

Zero-arg functions; repo root via __file__. Builds a tiny DATA.json, renders
both artifacts, and asserts: row IDs survive the round-trip, rollups are COMPUTED
(not stored) and correct, --check passes on freshly-rendered files and fails when
either artifact is mutated. Skips cleanly if openpyxl is unavailable.
"""
import importlib.util
import json
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent


def _load():
    spec = importlib.util.spec_from_file_location(
        "tasktracker_gen", REPO / "tools" / "tasktracker_gen.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _have_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _sample():
    return {
        "meta": {"title": "T", "live_on_stash": "9.9.9",
                 "generated_note": "note"},
        "columns": {"incomplete": [], "completed": []},
        "incomplete": [
            {"ID": "A1", "Category": "x", "Item / scope": "s1", "Status": "OPEN",
             "Tier": "A", "Sandbox-able?": "Yes", "Blocked by": "-",
             "Next action / notes": "n"},
            {"ID": "A2", "Category": "x", "Item / scope": "s2", "Status": "OPEN",
             "Tier": "B", "Sandbox-able?": "No", "Blocked by": "-",
             "Next action / notes": "n"},
            {"ID": "A3", "Category": "x", "Item / scope": "s3", "Status": "OPEN",
             "Tier": "A", "Sandbox-able?": "Yes", "Blocked by": "-",
             "Next action / notes": "n"},
        ],
        "completed": [
            {"ID": "C1", "Category": "x", "Item / scope": "s", "Version": "1.0",
             "State": "LIVE", "Notes": "ok"},
        ],
    }


def _write_data(d):
    with open(d / "TASK_TRACKER_DATA.json", "w", encoding="utf-8") as fh:
        json.dump(_sample(), fh, indent=2, ensure_ascii=False)


def test_rollups_are_computed_from_rows():
    m = _load()
    n_inc, n_comp, sandbox, tier = m._rollups(
        _sample()["incomplete"], _sample()["completed"])
    assert n_inc == 3 and n_comp == 1
    sb = dict(sandbox)
    assert sb["Yes"] == 2
    assert sb["No (operator/stash)"] == 1
    tr = dict(tier)
    assert tr["Tier A"] == 2 and tr["Tier B"] == 1


def test_md_totals_track_the_data_not_a_stored_string():
    m = _load()
    md = m.render_md(_sample())
    # totals line is derived: 3 incomplete / 1 completed, live string carried
    assert "3 incomplete · 1 completed" in md
    assert "Live on stash:** 9.9.9" in md
    assert "| A1 |" in md and "| C1 |" in md


def test_markdown_has_one_terminal_newline():
    md = _load().render_md(_sample())
    assert md.endswith("\n")
    assert not md.endswith("\n\n")


def test_render_then_check_is_in_sync():
    if not _have_openpyxl():
        return
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        _write_data(d)
        assert m.render(str(d)) == 0
        assert (d / "TASK_TRACKER.md").is_file()
        assert (d / "TASK_TRACKER.xlsx").is_file()
        assert m.check(str(d)) == 0  # freshly rendered -> IN-SYNC


def test_check_detects_md_drift():
    if not _have_openpyxl():
        return
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        _write_data(d)
        m.render(str(d))
        # mutate the md by hand -> drift
        p = d / "TASK_TRACKER.md"
        p.write_text(p.read_text(encoding="utf-8").replace("| A1 |", "| ZZ |"),
                     encoding="utf-8")
        assert m.check(str(d)) == 1


def test_check_detects_xlsx_drift():
    if not _have_openpyxl():
        return
    import openpyxl
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        _write_data(d)
        m.render(str(d))
        xp = d / "TASK_TRACKER.xlsx"
        wb = openpyxl.load_workbook(xp)
        wb["Incomplete"]["A2"] = "MUTATED"
        wb.save(xp)
        assert m.check(str(d)) == 1


def test_xlsx_signature_releases_file_handle():
    """Read-only signature workbooks must not block Windows cleanup."""
    if not _have_openpyxl():
        return
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        xp = Path(tmp) / "TASK_TRACKER.xlsx"
        m.render_xlsx(_sample(), xp)
        assert m._xlsx_signature(xp)
        xp.unlink()
        assert not xp.exists()


def test_gated_chain_renders_and_checks():
    if not _have_openpyxl():
        return
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        data = _sample()
        data["meta"]["gated_chain"] = ["step one", "step two"]
        data["meta"]["gated_chain_note"] = "ordered"
        (d / "TASK_TRACKER_DATA.json").write_text(
            json.dumps(data, ensure_ascii=False), encoding="utf-8")
        assert m.render(str(d)) == 0
        md = (d / "TASK_TRACKER.md").read_text(encoding="utf-8")
        assert "## Gated chain (sequencing)" in md
        assert "1. step one" in md and "2. step two" in md
        assert m.check(str(d)) == 0  # chain is part of the single source


def test_bootstrap_round_trips_through_xlsx():
    if not _have_openpyxl():
        return
    import openpyxl
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        # build an xlsx by hand, bootstrap to DATA.json, render, check
        wb = openpyxl.Workbook()
        wi = wb.active
        wi.title = "Incomplete"
        wi.append(m.INCOMPLETE_COLS)
        wi.append(["A1", "x", "s1", "OPEN", "A", "Yes", "-", "n"])
        wc = wb.create_sheet("Completed")
        wc.append(m.COMPLETED_COLS)
        wc.append(["C1", "x", "s", "1.0", "LIVE", "ok"])
        wb.save(d / "TASK_TRACKER.xlsx")
        assert m.bootstrap(str(d)) == 0
        data = json.loads((d / "TASK_TRACKER_DATA.json").read_text())
        assert [r["ID"] for r in data["incomplete"]] == ["A1"]
        assert [r["ID"] for r in data["completed"]] == ["C1"]
        assert m.render(str(d)) == 0
        assert m.check(str(d)) == 0


def _sample_with_awaiting():
    s = _sample()
    # an item whose remaining work is the operator's (built code, live-verify on
    # stash) — same schema as incomplete, but it belongs on its own page.
    s["awaiting_operator"] = [
        {"ID": "W1", "Category": "x", "Item / scope": "built; live-verify",
         "Status": "BUILT (live-verify pending)", "Tier": "A",
         "Sandbox-able?": "No", "Blocked by": "-", "Next action / notes": "operator"},
    ]
    return s


def test_awaiting_operator_sheet_renders_and_checks():
    """awaiting_operator (when present) renders a third md section + xlsx sheet,
    the totals count it, and render->check stays in sync."""
    if not _have_openpyxl():
        return
    import openpyxl
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        with open(d / "TASK_TRACKER_DATA.json", "w", encoding="utf-8") as fh:
            json.dump(_sample_with_awaiting(), fh, indent=2, ensure_ascii=False)
        assert m.render(str(d)) == 0
        md = (d / "TASK_TRACKER.md").read_text(encoding="utf-8")
        assert "Awaiting operator" in md, "md is missing the awaiting-operator section"
        assert "| W1 |" in md
        assert "awaiting operator" in md  # totals line counts it
        wb = openpyxl.load_workbook(d / "TASK_TRACKER.xlsx")
        assert "Awaiting operator" in wb.sheetnames, wb.sheetnames
        assert wb["Awaiting operator"]["A2"].value == "W1"
        assert m.check(str(d)) == 0  # freshly rendered -> IN-SYNC


def test_awaiting_operator_absent_is_backward_compatible():
    """With no awaiting_operator key the workbook is the prior two data sheets
    only — the page is strictly additive."""
    if not _have_openpyxl():
        return
    import openpyxl
    m = _load()
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        _write_data(d)  # _sample() has no awaiting_operator key
        assert m.render(str(d)) == 0
        wb = openpyxl.load_workbook(d / "TASK_TRACKER.xlsx")
        assert "Awaiting operator" not in wb.sheetnames
        md = (d / "TASK_TRACKER.md").read_text(encoding="utf-8")
        assert "Awaiting operator" not in md
        assert m.check(str(d)) == 0
