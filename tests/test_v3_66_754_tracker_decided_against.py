"""v3.66.754a -- `decided_against` was a WRITE-ONLY section of the ledger.

THE DEFECT. `tasktracker_gen.py` hardcoded its audit denominator:

    sections = [("incomplete", ...), ("awaiting_operator", ...), ("completed", ...)]

`decided_against` is a FOURTH declared section, and it appeared in none of the three
consumers -- not the auditor, not the markdown renderer, not the xlsx renderer. So it was
**neither audited, nor rendered, nor counted**: 16 rows of DECISIONS existed only in the
JSON, visible to nothing that a session actually reads.

Proof of the audit half (the one that matters): inject a row with a garbage schema into
`decided_against` and `--audit` returns **"[OK] schema conforms | RESULT: CLEAN"**. The
check is truthful and useless -- its denominator structurally excludes the thing being
asked about. That is the dominant failure shape of this entire program, sitting inside the
ledger's own gate.

WHY IT MATTERS OPERATIONALLY. The corrective discipline says: *your CLOSED verdicts belong
in `decided_against` -- the reason IS the artifact.* So the thumbs/a11y closure reasons were
filed into a section no reader ever sees. A future session finds three dark `thumbs/*`
endpoints, finds no record of why they were closed, and wires an arbitrary-file-read surface
into the GUI. The write-only section is not a cosmetic gap; it is how a security decision
gets silently un-made.

HISTORY. Fixed at 753 in the PK copy ONLY. The repo copy at `tools/tasktracker_gen.py` --
the one a session actually runs from the work tree, and the one `build_session_pack.py`
imports -- stayed blind. The next pack built from a pristine tree would have resurrected the
blackout. That is `FG-PK-MIRROR-DRIFTS-SILENTLY` / KB_JUDGMENT (j): the same document in two
channels, and only one got fixed. Tracked as TRACKER-GEN-TREE-COPY (P1).

These tests pin all three consumers. RED-first: every one fails on the pristine tree.
"""
import importlib.util
import json
import os
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TOOL = os.path.join(ROOT, "tools", "tasktracker_gen.py")


def _mod():
    spec = importlib.util.spec_from_file_location("tasktracker_gen", TOOL)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


def _write(d, data):
    """Use the tool's own DATA_NAME -- _load_data sys.exit()s on a miss, which would
    surface as a test failure rather than a bad fixture."""
    m = _mod()
    os.makedirs(d, exist_ok=True)
    json.dump(data, open(os.path.join(d, m.DATA_NAME), "w"))


# The decided_against contract, stated INDEPENDENTLY of the implementation. Deriving the
# fixture from m.DECIDED_COLS would make every test die in setup when the constant is
# absent (a RED for the wrong reason), and would mirror the very thing under test.
DECIDED_CONTRACT = ["ID", "Category", "Item / scope", "Version", "State", "Notes"]


def _have_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _decided(**over):
    row = {c: "x" for c in DECIDED_CONTRACT}
    row["ID"] = "DA-1"
    row.update(over)
    return row


# --------------------------------------------------------------------------
# 1. The schema constant must exist at all.
# --------------------------------------------------------------------------
def test_decided_cols_is_declared():
    m = _mod()
    assert hasattr(m, "DECIDED_COLS"), (
        "DECIDED_COLS is not declared -- the decided_against section has no schema, so "
        "nothing can audit or render it")
    assert list(m.DECIDED_COLS) == DECIDED_CONTRACT, (
        "DECIDED_COLS does not match the shipped decided_against schema %r" % (DECIDED_CONTRACT,))


# --------------------------------------------------------------------------
# 2. THE ONE THAT MATTERS: the audit's denominator must CONTAIN the section.
#    A check that cannot see its subject reports clean -- truthfully, uselessly.
# --------------------------------------------------------------------------
def test_audit_catches_a_garbage_decided_against_row():
    m = _mod()
    with tempfile.TemporaryDirectory() as d:
        garbage = {"wrong": "schema", "totally": "bogus"}   # conforms to nothing
        _write(d, {"incomplete": [], "completed": [],
                   "decided_against": [garbage]})
        rc = m.audit(d)
    assert rc == 1, (
        "audit() returned CLEAN over a decided_against row that conforms to no schema. "
        "The section is not in the audit's denominator, so the gate is green because it "
        "cannot see the thing it is being asked about -- not because the thing is sound.")


def test_audit_passes_a_well_formed_decided_against_row():
    """The POS half: adding the section to the denominator must not make a VALID
    ledger fail. A gate that fires on everything is as useless as one that fires on
    nothing."""
    m = _mod()
    with tempfile.TemporaryDirectory() as d:
        _write(d, {"incomplete": [], "completed": [],
                   "decided_against": [_decided()]})
        rc = m.audit(d)
    assert rc == 0, "a well-formed decided_against row must not trip the audit"


def test_audit_catches_a_duplicate_id_in_decided_against():
    """Dup-ID detection sweeps `sections`, so a section outside the denominator is also
    invisible to it: an item parked in BOTH incomplete and decided_against reads clean."""
    m = _mod()
    with tempfile.TemporaryDirectory() as d:
        inc = {c: "x" for c in m.INCOMPLETE_COLS}
        inc["ID"] = "DUP-1"
        _write(d, {"incomplete": [inc], "completed": [],
                   "decided_against": [_decided(ID="DUP-1")]})
        rc = m.audit(d)
    assert rc == 1, (
        "an ID present in BOTH incomplete and decided_against was not flagged -- the "
        "dup-ID sweep iterates `sections`, so it inherits the same blind denominator")


# --------------------------------------------------------------------------
# 3. Rendering. Auditing it but not showing it would still leave it unreadable.
#    (Fixing render_md alone would have left the xlsx blind -- pin BOTH.)
# --------------------------------------------------------------------------
def test_render_md_emits_decided_against():
    m = _mod()
    row = _decided(ID="DA-VISIBLE")
    md = m.render_md({"incomplete": [], "completed": [], "decided_against": [row]})
    assert "DA-VISIBLE" in md, (
        "render_md drops decided_against -- the reasons for every CLOSED verdict are "
        "invisible in the document a session actually reads")


def test_render_xlsx_emits_decided_against():
    if not _have_openpyxl():
        return  # openpyxl absent -- the md half still pins the renderer contract
    import openpyxl
    m = _mod()
    row = _decided(ID="DA-XLSX")
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "TASK_TRACKER.xlsx")
        m.render_xlsx({"incomplete": [], "completed": [], "decided_against": [row]}, p)
        wb = openpyxl.load_workbook(p)
        found = any(
            any(str(c.value) == "DA-XLSX" for r in wb[s].iter_rows() for c in r)
            for s in wb.sheetnames)
    assert found, (
        "render_xlsx drops decided_against -- fixing the markdown renderer alone would "
        "have left the spreadsheet blind to the same rows")


# --------------------------------------------------------------------------
# 4. Anti-regression on the SHAPE, not just this instance: the audit must derive its
#    sections from what the DATA declares, so the next section added is not invisible
#    by default. This is the actual lesson -- a hardcoded denominator will drift again.
# --------------------------------------------------------------------------
def test_audit_counts_report_the_decided_against_section():
    m = _mod(); import io, contextlib
    with tempfile.TemporaryDirectory() as d:
        _write(d, {"incomplete": [], "completed": [],
                   "decided_against": [_decided(), _decided(ID="DA-2")]})
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            m.audit(d)
        out = buf.getvalue()
    assert "decided_against" in out, (
        "the audit's own count line omits decided_against -- the rows are not COUNTED, "
        "so a section could empty out silently and the summary would still read clean")


# ==========================================================================
# 5. THE FOURTH CONSUMER -- tasktracker_sync, the IN-SYNC GATE.
#
# The 753 fix covered the three consumers inside tasktracker_gen (audit,
# render_md, render_xlsx) and stopped there. `tasktracker_sync` is a FOURTH
# consumer with the SAME hardcoded denominator, and it is the one that
# `build_session_pack.py:114` HARD-GATES the pack build on:
#
#     if tts.check(a.pack_dir) != 0:
#         sys.exit("FAIL: TASK_TRACKER xlsx<->md DRIFT (fix before packing)")
#
# Two independent blindnesses, measured on the shipped tracker:
#   * xlsx_ids reads only the "Incomplete" and "Completed" SHEETS -- the workbook
#     actually ships five ('Summary', 'Incomplete', 'Awaiting operator',
#     'Completed', 'Decided against').
#   * md_ids is a state machine with NO RESET on an unrecognised heading: it sets
#     sheet="Completed" at "## Completed" and then appends EVERY table row to EOF,
#     swallowing the "## Decided against" table. Same for awaiting -> Incomplete.
#
# Measured drift on the real shipped pack: Incomplete md=38 vs xlsx=12 (26 awaiting
# rows swallowed); Completed md=374 vs xlsx=359 (16 decided rows swallowed). The gate
# was comparing two differently-wrong numbers, so it could report DRIFT for a tracker
# that is perfectly in sync -- and would happily pass a real drift that happened to
# cancel out. A gate whose two sides are both wrong is not a gate.
#
# This is the SAME shape as the audit defect above, in a sibling file, which is why
# fixing only tasktracker_gen would have left the pack build broken.
# ==========================================================================
import importlib.util as _ilu

SYNC = os.path.join(ROOT, "tools", "tasktracker_sync.py")


def _sync():
    spec = _ilu.spec_from_file_location("tasktracker_sync", SYNC)
    m = _ilu.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


def _full_tracker():
    """A tracker with all four sections populated and DISTINCT ids per section."""
    g = _mod()
    inc = {c: "x" for c in g.INCOMPLETE_COLS}; inc["ID"] = "INC-1"
    awa = {c: "x" for c in g.AWAITING_COLS};   awa["ID"] = "AWA-1"
    comp = {c: "x" for c in g.COMPLETED_COLS}; comp["ID"] = "CMP-1"
    dec = _decided(ID="DEC-1")
    return {"incomplete": [inc], "awaiting_operator": [awa],
            "completed": [comp], "decided_against": [dec]}


def test_sync_md_parser_does_not_swallow_later_sections():
    """The md parser must STOP at an unrecognised heading, not keep appending. A
    decided_against row landing in the Completed bucket is how the gate ended up
    comparing two differently-wrong numbers."""
    g, s = _mod(), _sync()
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "TASK_TRACKER.md")
        open(p, "w", encoding="utf-8").write(g.render_md(_full_tracker()))
        ids = s.md_ids(p)
    assert "DEC-1" not in ids.get("Completed", []), (
        "md_ids swallowed a decided_against row into Completed -- the state machine "
        "never resets on a heading it does not recognise")
    assert "AWA-1" not in ids.get("Incomplete", []), (
        "md_ids swallowed an awaiting_operator row into Incomplete -- same defect")


def test_sync_xlsx_reader_sees_every_rendered_sheet():
    """xlsx_ids read 2 of the 5 sheets the renderer actually emits."""
    if not _have_openpyxl():
        return
    g, s = _mod(), _sync()
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "TASK_TRACKER.xlsx")
        g.render_xlsx(_full_tracker(), p)
        ids = s.xlsx_ids(p)
    flat = [i for v in ids.values() for i in v]
    for want in ("INC-1", "AWA-1", "CMP-1", "DEC-1"):
        assert want in flat, (
            "xlsx_ids does not read the sheet containing %r -- it inspects only a "
            "hardcoded subset of the sheets the renderer emits" % want)


def test_sync_check_is_green_on_a_freshly_rendered_tracker():
    """THE ONE THAT GATES THE PACK BUILD. Render md+xlsx from the SAME data with the
    project's own renderers, then ask the in-sync gate. It must say IN-SYNC. If the two
    readers disagree about what they are counting, this fails -- which is exactly the
    state that blocks `build_session_pack`."""
    if not _have_openpyxl():
        return
    g, s = _mod(), _sync()
    data = _full_tracker()
    with tempfile.TemporaryDirectory() as d:
        json.dump(data, open(os.path.join(d, g.DATA_NAME), "w"))
        open(os.path.join(d, "TASK_TRACKER.md"), "w", encoding="utf-8").write(g.render_md(data))
        g.render_xlsx(data, os.path.join(d, "TASK_TRACKER.xlsx"))
        rc = s.check(d)
    assert rc == 0, (
        "tasktracker_sync.check reports DRIFT on artifacts BOTH rendered from the same "
        "data by the project's own renderers. The gate build_session_pack.py:114 blocks "
        "the pack build on is comparing two differently-wrong numbers.")


def test_sync_md_parser_does_not_mistake_a_data_row_for_the_header():
    """The header-skip was `not ln.startswith("| ID")` -- a PREFIX match. So the real
    completed row `| IDEA-HARDEN | ...` was mistaken for the header row and SILENTLY
    DROPPED (measured: md=358 vs xlsx=359 the moment the section denominator was fixed).

    It survived because BOTH sides of the comparison were wrong and the errors cancelled:
    md was over-counting by swallowing 16 decided rows and under-counting by 1 here. This
    is the precise hazard the docstring above predicted -- a gate with two wrong sides can
    pass a real drift that happens to cancel out. Match the header CELL, not a prefix."""
    g, s = _mod(), _sync()
    data = _full_tracker()
    comp = {c: "x" for c in g.COMPLETED_COLS}
    comp["ID"] = "IDEA-HARDEN"          # a legitimate ID that begins with "ID"
    data["completed"].append(comp)
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "TASK_TRACKER.md")
        open(p, "w", encoding="utf-8").write(g.render_md(data))
        ids = s.md_ids(p)
    assert "IDEA-HARDEN" in ids["Completed"], (
        "md_ids dropped a real row whose ID starts with 'ID' -- the header-skip is a "
        "prefix match, so any such row is silently mistaken for the header")
    assert "ID" not in ids["Completed"], "the actual header row must still be skipped"
