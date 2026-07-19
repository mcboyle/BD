#!/usr/bin/env python3
"""tasktracker_gen — render TASK_TRACKER.md AND TASK_TRACKER.xlsx from ONE
canonical Python/JSON data source (``TASK_TRACKER_DATA.json``).

Background. The xlsx was the source and the md a hand-or-tool mirror
(``tasktracker_sync.py`` regenerates md from xlsx). But the *xlsx itself* was
hand-edited, and the md preamble totals were typed by hand — so the table IDs
could be IN-SYNC while the prose totals said "51 · 69" against a real 48/73, and
"live 3.66.276" long after 285 deployed. This tool closes that last gap: a single
JSON source, with every rollup (totals, by Sandbox-able?, by Tier) *computed* at
render time, emitted identically into both artifacts. Prose drift becomes
structurally impossible.

  python3 tools/tasktracker_gen.py --bootstrap <dir>   # one-time: xlsx -> DATA.json
  python3 tools/tasktracker_gen.py --render    <dir>   # DATA.json -> .md + .xlsx
  python3 tools/tasktracker_gen.py --check     <dir>   # exit 1 if either artifact drifts
  python3 tools/tasktracker_gen.py --render            # cwd

DATA.json shape (the canonical source):
  {
    "meta": {"title","generated_note","live_on_stash","gate_line","priority_note"},
    "columns": {"incomplete":[...8 headers...], "completed":[...6 headers...]},
    "incomplete": [ {<header>: <value>, ...}, ... ],
    "completed":  [ {<header>: <value>, ...}, ... ]
  }

Rollups are NOT stored — they are derived from the ``incomplete`` rows so they
can never disagree with the table. ``--render`` rewrites both files; ``--check``
renders to a temp dir and diffs, exiting 1 (with the offending artifact named) on
any difference. Supersedes ``tasktracker_sync.py`` (which remains as the legacy
xlsx->md mirror + its pinned test); the canonical gate is now ``--check`` here.

Requires openpyxl (sandbox + service venv) — a dev/close-time tool, not a
stash-runtime chain CLI, so the non-stdlib dep is acceptable. Fails loudly if
openpyxl is missing.
"""
import argparse
import json
import os
import sys
import tempfile

DATA_NAME = "TASK_TRACKER_DATA.json"
MD_NAME = "TASK_TRACKER.md"
XLSX_NAME = "TASK_TRACKER.xlsx"

INCOMPLETE_COLS = ["ID", "Category", "Item / scope", "Status", "Tier",
                   "Sandbox-able?", "Blocked by", "Next action / notes"]
COMPLETED_COLS = ["ID", "Category", "Item / scope", "Version", "State", "Notes"]
# Awaiting-operator items reuse the incomplete schema verbatim: they ARE
# incomplete (work remains), but the remaining work is the operator's hands on
# stash (live-verify of built code, operator-run verification, operator-only
# tasks) rather than sandbox/dev work. Same columns -> a moved row keeps every
# field unchanged. Optional key: absent/empty -> rendered output is byte-for-byte
# the prior two-sheet workbook.
AWAITING_COLS = INCOMPLETE_COLS
# v3.66.754a: decided_against is a FOURTH declared section and it appeared in NONE of the
# three consumers -- not the auditor, not render_md, not render_xlsx. 16 rows of decisions
# existed only in the JSON. It reused the completed schema in shape but had no constant of
# its own, so nothing could audit or render it.
DECIDED_COLS = ["ID", "Category", "Item / scope", "Version", "State", "Notes"]

# The section registry. The audit used to carry a HARDCODED list of three sections, so a
# fourth was invisible BY CONSTRUCTION -- the gate reported clean over a denominator that
# structurally excluded the thing being asked about. Consumers now derive their sections
# from what the DATA DECLARES against this registry, and an UNRECOGNISED section is a
# LOUD failure rather than a silent skip: that is what stops the next section added from
# being invisible by default. Fixing only decided_against would have left the shape intact.
SECTION_SCHEMAS = {
    "incomplete": INCOMPLETE_COLS,
    "awaiting_operator": AWAITING_COLS,
    "completed": COMPLETED_COLS,
    "decided_against": DECIDED_COLS,
}
SECTION_ORDER = ["incomplete", "awaiting_operator", "completed", "decided_against"]
# keys in DATA.json that are metadata, not row sections
_NON_SECTION_KEYS = {"meta", "columns"}


# --------------------------------------------------------------------------- #
# rollups (derived, never stored)
# --------------------------------------------------------------------------- #
def _rollups(incomplete, completed):
    """Compute the Summary rollups from the row data alone."""
    n_inc, n_comp = len(incomplete), len(completed)

    def _bucket(rows, key, order, fallback):
        counts = {}
        for r in rows:
            v = (r.get(key) or "").strip()
            counts[v if v else fallback] = counts.get(v if v else fallback, 0) + 1
        # emit in declared order first, then any leftovers alphabetically
        out = [(label, counts[raw]) for label, raw in order if raw in counts]
        seen = {raw for _, raw in order}
        for raw in sorted(k for k in counts if k not in seen):
            out.append((raw, counts[raw]))
        return out

    sandbox = _bucket(
        incomplete, "Sandbox-able?",
        [("Yes", "Yes"), ("Partial", "Partial"), ("No (operator/stash)", "No")],
        "(unset)")
    # the legacy Summary labels "No" as "No (operator/stash)"; normalize display
    sandbox = [("No (operator/stash)" if lbl == "No" else lbl, n) for lbl, n in sandbox]

    tier = _bucket(
        incomplete, "Tier",
        [("Tier A", "A"), ("Tier B", "B"), ("Tier M", "M"),
         ("Tier S", "S"), ("Tier L", "L"), ("Tier -", "-")],
        "(unset)")
    return n_inc, n_comp, sandbox, tier


# --------------------------------------------------------------------------- #
# render
# --------------------------------------------------------------------------- #
def _md_escape(s):
    return str(s).replace("|", "\\|").replace("\n", " ")


def _md_table(cols, rows):
    out = ["| " + " | ".join(cols) + " |",
           "| " + " | ".join("---" for _ in cols) + " |"]
    for r in rows:
        out.append("| " + " | ".join(_md_escape(r.get(c, "")) for c in cols) + " |")
    return "\n".join(out)


def render_md(data):
    meta = data.get("meta", {})
    inc, comp = data["incomplete"], data["completed"]
    awaiting = data.get("awaiting_operator", []) or []
    decided = data.get("decided_against", []) or []
    n_inc, n_comp, sandbox, tier = _rollups(inc, comp)
    title = meta.get("title", "BulkDownloader — unified task tracker")
    parts = [f"# {title}", ""]
    if meta.get("generated_note"):
        parts += [f"*{meta['generated_note']}*", ""]
    totals = f"**Totals:** {n_inc} incomplete · {n_comp} completed."
    if awaiting:
        totals = (f"**Totals:** {n_inc} incomplete · {n_comp} completed · "
                  f"{len(awaiting)} awaiting operator.")
    if meta.get("live_on_stash"):
        gate = f" ({meta['gate_line']})" if meta.get("gate_line") else ""
        totals += f" **Live on stash:** {meta['live_on_stash']}{gate}."
    parts += [totals, ""]
    if meta.get("priority_note"):
        parts += [f"**Priority:** {meta['priority_note']}", ""]
    chain = meta.get("gated_chain")
    if chain:
        parts += ["## Gated chain (sequencing)", ""]
        if meta.get("gated_chain_note"):
            parts += [meta["gated_chain_note"], ""]
        for i, step in enumerate(chain, 1):
            parts.append(f"{i}. {step}")
        parts.append("")
    parts += ["## Incomplete (running)", "", _md_table(INCOMPLETE_COLS, inc), ""]
    if awaiting:
        parts += ["## Awaiting operator (built — live-verify or operator action on stash)",
                  "", _md_table(AWAITING_COLS, awaiting), ""]
    parts += ["## Completed (running)", "", _md_table(COMPLETED_COLS, comp), ""]
    if decided:
        parts += ["## Decided against (closed by decision — the reason IS the artifact)",
                  "", _md_table(DECIDED_COLS, decided), ""]
    return "\n".join(parts) + "\n"


def render_xlsx(data, path):
    """Render a sleek, human-readable workbook.

    IMPORTANT — drift contract: ``--check`` compares workbooks by a
    *value-only* signature (``_xlsx_signature``: ``data_only`` + ``values_only``),
    so ALL styling here is invisible to the gate **provided cell VALUES and the
    sheet SHAPE (row/column extent + sheet set) are unchanged**. This function
    therefore styles freely but never adds/removes a row, column, or sheet, and
    never merges past an existing column extent — so the rendered file stays
    IN-SYNC under both this generator and the prior (plain) one. Sheet *order* is
    cosmetic (the signature is keyed by sheet name), so Summary leads.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("tasktracker_gen: openpyxl required (run in venv/sandbox): "
                 "pip install openpyxl --break-system-packages")
    meta = data.get("meta", {})
    inc, comp = data["incomplete"], data["completed"]
    awaiting = data.get("awaiting_operator", []) or []
    decided = data.get("decided_against", []) or []
    n_inc, n_comp, sandbox, tier = _rollups(inc, comp)

    # ---- palette ---------------------------------------------------------- #
    INK = "1F2A44"
    HEAD_FILL = PatternFill("solid", fgColor=INK)
    HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEAD_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)
    HEAD_BORDER = Border(bottom=Side(style="medium", color=INK))
    ID_FONT = Font(bold=True, color=INK)
    TITLE_FONT = Font(bold=True, size=15, color=INK)
    NOTE_FONT = Font(italic=True, size=9, color="6B7280")
    SECT_FONT = Font(bold=True, size=11, color=INK)
    SECT_FILL = PatternFill("solid", fgColor="E8ECF4")
    ZEBRA = PatternFill("solid", fgColor="F5F7FB")
    WRAP = Alignment(vertical="top", wrap_text=True)
    CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)
    NUM_ALIGN = Alignment(vertical="center", horizontal="right")
    ROW_BORDER = Border(bottom=Side(style="thin", color="E5E9F0"))

    def chip(bg, fg):
        return PatternFill("solid", fgColor=bg), Font(color=fg, bold=True)
    GREEN, AMBER, RED, GREY = (chip("DCFCE7", "166534"), chip("FEF3C7", "92400E"),
                               chip("FEE2E2", "991B1B"), chip("EEF1F6", "475569"))
    TIER_TINT = {"A": "FFE4E6", "B": "E0F2FE", "M": "F1F5F9",
                 "S": "FAE8FF", "L": "ECFCCB", "-": "F8FAFC"}

    def state_chip(v):
        t = (v or "").strip().lower()
        if not t:
            return None
        if any(k in t for k in ("live", "done", "complete", "shipped", "merged")):
            return GREEN
        if any(k in t for k in ("built", "pending", "wip", "draft", "progress")):
            return AMBER
        return GREY

    def status_chip(v):
        t = (v or "").lower()
        if "block" in t:
            return RED
        if any(k in t for k in ("pending", "await", "deferred", "parked", "tbd")):
            return AMBER
        return None

    def sandbox_chip(v):
        t = (v or "").strip().lower()
        if t.startswith("yes"):
            return GREEN
        if t.startswith("no"):
            return RED
        if "partial" in t:
            return AMBER
        return None

    def _table(ws, cols, rows, widths, tab):
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = tab
        ws.append(cols)
        for c in ws[1]:
            c.font, c.fill, c.alignment, c.border = (
                HEAD_FONT, HEAD_FILL, HEAD_ALIGN, HEAD_BORDER)
        ws.row_dimensions[1].height = 30
        for i, r in enumerate(rows, start=2):
            ws.append([r.get(col, "") for col in cols])
            zebra = (i % 2 == 0)
            for j, col in enumerate(cols, start=1):
                cell = ws.cell(row=i, column=j)
                cell.alignment = WRAP
                cell.border = ROW_BORDER
                if zebra:
                    cell.fill = ZEBRA
                val = r.get(col, "")
                if col == "ID":
                    cell.font = ID_FONT
                elif col == "Tier":
                    cell.alignment = CENTER
                    cell.font = Font(bold=True, color=INK)
                    tint = TIER_TINT.get((val or "").strip())
                    if tint:
                        cell.fill = PatternFill("solid", fgColor=tint)
                elif col == "Sandbox-able?":
                    cell.alignment = CENTER
                    ch = sandbox_chip(val)
                    if ch:
                        cell.fill, cell.font = ch
                elif col == "Status":
                    ch = status_chip(val)
                    if ch:
                        cell.fill, cell.font = ch
                elif col == "State":
                    cell.alignment = CENTER
                    ch = state_chip(val)
                    if ch:
                        cell.fill, cell.font = ch
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "B2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
        return ws

    wb = openpyxl.Workbook()

    # ---- Summary (leads) -------------------------------------------------- #
    ws_s = wb.active
    ws_s.title = "Summary"
    ws_s.sheet_view.showGridLines = False
    ws_s.sheet_properties.tabColor = "3B82F6"

    def srow(values):
        ws_s.append(values)
        return ws_s.max_row

    def section(label):
        rn = srow([label])
        ws_s.cell(row=rn, column=1).font = SECT_FONT
        ws_s.cell(row=rn, column=1).fill = SECT_FILL
        ws_s.cell(row=rn, column=2).fill = SECT_FILL
        return rn

    r = srow([meta.get("title", "BulkDownloader — unified task tracker")])
    ws_s.cell(row=r, column=1).font = TITLE_FONT
    ws_s.row_dimensions[r].height = 22
    if meta.get("generated_note"):
        r = srow([meta["generated_note"]])
        ws_s.cell(row=r, column=1).font = NOTE_FONT
        ws_s.cell(row=r, column=1).alignment = WRAP
    srow([])
    section("Totals")
    _total_rows = [("Incomplete items", n_inc)]
    if awaiting:
        _total_rows.append(("Awaiting operator", len(awaiting)))
    _total_rows.append(("Completed items", n_comp))
    for lbl, n in _total_rows:
        r = srow([lbl, n])
        ws_s.cell(row=r, column=2).font = Font(bold=True)
        ws_s.cell(row=r, column=2).alignment = NUM_ALIGN
    srow([])
    section("Incomplete by Sandbox-able?")
    for lbl, n in sandbox:
        r = srow([lbl, n])
        ws_s.cell(row=r, column=2).alignment = NUM_ALIGN
    srow([])
    section("Incomplete by Tier")
    for lbl, n in tier:
        r = srow([lbl, n])
        ws_s.cell(row=r, column=2).alignment = NUM_ALIGN
    chain = meta.get("gated_chain")
    if chain:
        srow([])
        section("Gated chain (sequencing)")
        for i, step in enumerate(chain, 1):
            r = srow([str(i), step])
            ws_s.cell(row=r, column=1).alignment = CENTER
            ws_s.cell(row=r, column=1).font = Font(bold=True, color=INK)
            ws_s.cell(row=r, column=2).alignment = WRAP
    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 96

    # ---- data sheets ------------------------------------------------------ #
    _table(wb.create_sheet("Incomplete"), INCOMPLETE_COLS, inc,
           [9, 22, 58, 22, 7, 13, 16, 58], "F59E0B")
    if awaiting:
        _table(wb.create_sheet("Awaiting operator"), AWAITING_COLS, awaiting,
               [9, 22, 58, 22, 7, 13, 16, 58], "8B5CF6")
    _table(wb.create_sheet("Completed"), COMPLETED_COLS, comp,
           [10, 22, 48, 14, 11, 72], "22C55E")
    if decided:
        _table(wb.create_sheet("Decided against"), DECIDED_COLS, decided,
               [10, 22, 48, 14, 11, 72], "64748B")

    wb.active = 0
    wb.save(path)


def render(d):
    data = _load_data(d)
    md = render_md(data)
    with open(os.path.join(d, MD_NAME), "w", encoding="utf-8") as fh:
        fh.write(md)
    render_xlsx(data, os.path.join(d, XLSX_NAME))
    n_inc, n_comp, _, _ = _rollups(data["incomplete"], data["completed"])
    print(f"  rendered {MD_NAME} + {XLSX_NAME}: "
          f"{n_inc} incomplete, {n_comp} completed (rollups computed)")
    return 0


# --------------------------------------------------------------------------- #
# check (render to temp, diff)
# --------------------------------------------------------------------------- #
def _xlsx_signature(path):
    """A comparable, style-independent view of the workbook: per-sheet header +
    cell-value rows. Avoids brittle byte-compare (zip mtimes/ordering)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sig = {}
    for s in wb.sheetnames:
        sig[s] = [tuple("" if c is None else str(c) for c in row)
                  for row in wb[s].iter_rows(values_only=True)]
    return sig


def check(d):
    if not os.path.isfile(os.path.join(d, DATA_NAME)):
        sys.exit(f"tasktracker_gen: missing {os.path.join(d, DATA_NAME)} "
                 f"(run --bootstrap first)")
    drift = False
    # MD: text compare
    want_md = render_md(_load_data(d))
    mp = os.path.join(d, MD_NAME)
    have_md = open(mp, encoding="utf-8").read() if os.path.isfile(mp) else None
    if have_md != want_md:
        drift = True
        print(f"  [DRIFT] {MD_NAME} differs from {DATA_NAME}")
    else:
        print(f"  [OK] {MD_NAME}")
    # XLSX: value-signature compare (render to temp)
    xp = os.path.join(d, XLSX_NAME)
    with tempfile.TemporaryDirectory() as tmp:
        ref = os.path.join(tmp, XLSX_NAME)
        render_xlsx(_load_data(d), ref)
        if not os.path.isfile(xp):
            drift = True
            print(f"  [DRIFT] {XLSX_NAME} missing")
        elif _xlsx_signature(xp) != _xlsx_signature(ref):
            drift = True
            print(f"  [DRIFT] {XLSX_NAME} differs from {DATA_NAME}")
        else:
            print(f"  [OK] {XLSX_NAME}")
    print("RESULT:", "DRIFT" if drift else "IN-SYNC")
    return 1 if drift else 0


# --------------------------------------------------------------------------- #
# bootstrap (xlsx -> DATA.json), one-time
# --------------------------------------------------------------------------- #
def bootstrap(d):
    try:
        import openpyxl
    except ImportError:
        sys.exit("tasktracker_gen: openpyxl required for --bootstrap")
    xp = os.path.join(d, XLSX_NAME)
    if not os.path.isfile(xp):
        sys.exit(f"tasktracker_gen: missing {xp}")
    wb = openpyxl.load_workbook(xp, data_only=True)

    def _rows(sheet, cols):
        out = []
        if sheet not in wb.sheetnames:
            return out
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            out.append({c: ("" if (i >= len(row) or row[i] is None) else str(row[i]))
                        for i, c in enumerate(cols)})
        return out

    # preamble + totals line live in the Summary sheet's first rows
    meta = {"title": "BulkDownloader — unified task tracker"}
    if "Summary" in wb.sheetnames:
        srows = list(wb["Summary"].iter_rows(values_only=True))
        if len(srows) > 1 and srows[1] and srows[1][0]:
            meta["generated_note"] = str(srows[1][0])
    data = {
        "meta": meta,
        "columns": {"incomplete": INCOMPLETE_COLS, "completed": COMPLETED_COLS},
        "incomplete": _rows("Incomplete", INCOMPLETE_COLS),
        "completed": _rows("Completed", COMPLETED_COLS),
    }
    with open(os.path.join(d, DATA_NAME), "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
        fh.write("\n")
    print(f"  bootstrapped {DATA_NAME}: {len(data['incomplete'])} incomplete, "
          f"{len(data['completed'])} completed (review meta.* before --render)")
    return 0


def _load_data(d):
    p = os.path.join(d, DATA_NAME)
    if not os.path.isfile(p):
        sys.exit(f"tasktracker_gen: missing {p} (run --bootstrap first)")
    with open(p, encoding="utf-8") as fh:
        data = json.load(fh)
    data.setdefault("incomplete", [])
    data.setdefault("completed", [])
    data.setdefault("meta", {})
    return data


def status(d):
    """Print a human-readable rollup of the tracker counts from DATA.json.

    stdlib-only (no openpyxl) so a fresh sandbox session can call it from
    bd-boot / bd-status. DEGRADES GRACEFULLY -- a clear, actionable message
    and exit 0 (never a traceback) -- when TASK_TRACKER_DATA.json is absent
    (e.g. it was not bundled into version.zip): counts genuinely cannot be
    derived without the data file, so we say so rather than fabricating
    numbers or crashing. Rollups are computed (never stored), mirroring the
    Summary sheet's by-Sandbox-able? / by-Tier breakdown.
    """
    p = os.path.join(d, DATA_NAME)
    if not os.path.exists(p):
        print(f"  [no data] {DATA_NAME} not found in {os.path.abspath(d)}")
        print(f"  counts unavailable -- bundle {DATA_NAME} into version.zip, "
              f"or run 'tasktracker_gen.py --status <dir>' on stash where the "
              f"data file lives.")
        return 0
    try:
        with open(p, encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception as e:  # unreadable/corrupt -> graceful, still no crash
        print(f"  [no data] {DATA_NAME} present but unreadable: {str(e)[:120]}")
        print(f"  counts unavailable -- regenerate it with "
              f"'tasktracker_gen.py --bootstrap <dir>' on stash.")
        return 0
    inc = data.get("incomplete", []) or []
    comp = data.get("completed", []) or []
    awaiting = data.get("awaiting_operator", []) or []
    total = len(inc) + len(comp) + len(awaiting)
    n_inc, n_comp, sandbox, tier = _rollups(inc, comp)
    print(f"  TASK TRACKER STATUS ({DATA_NAME})")
    print(f"    incomplete:        {len(inc)}")
    print(f"    awaiting-operator: {len(awaiting)}")
    print(f"    completed:         {len(comp)}")
    print(f"    total:             {total}")
    if sandbox:
        print("    incomplete by Sandbox-able?: "
              + ", ".join(f"{lbl} {n}" for lbl, n in sandbox))
    if tier:
        print("    incomplete by Tier: "
              + ", ".join(f"{lbl} {n}" for lbl, n in tier))
    return 0


# v3.66.721: RESTORED. This function was DROPPED from the work tree (PK still had
# it; the same edit appears to have added status()). It is the only one of the nine
# PK/work-tree divergences where the work tree LOST something -- the other eight are
# the work tree being newer. --check only proves the md and xlsx agree with EACH
# OTHER, so it is green whenever they agree, including when they agree on garbage.
# --audit checks the DATA against its SCHEMA, which is the question actually asked.
def audit(d):
    """Structural audit that ``--check`` cannot do: ``--check`` only proves the
    md and xlsx agree with each other (render drift). It passed IN-SYNC this
    release while a completed-section row carried the *incomplete* schema and
    rendered blank Version/State/Notes. ``--audit`` catches that class:

      1. per-section SCHEMA conformance (incomplete/awaiting use INCOMPLETE_COLS;
         completed uses COMPLETED_COLS) -- the P4-GAPS defect,
      2. DUPLICATE IDs across sections (an item left in two places),
      3. STATE.json prose vs DATA.json counts (if STATE.json is alongside) --
         the "AS OF 328" drift, where the narrative lags the data.

    Exit 1 on any defect, 0 clean. Pairs with --check (render) for a full gate.
    """
    import re
    from collections import Counter
    data = _load_data(d)
    # Derive the denominator from what the data DECLARES (see SECTION_SCHEMAS). The old
    # hardcoded 3-tuple could not see decided_against, so a row conforming to no schema at
    # all returned "RESULT: CLEAN".
    sections = [(s_, set(SECTION_SCHEMAS[s_])) for s_ in SECTION_ORDER
                if s_ in data or s_ in ("incomplete", "completed")]

    problems = []
    # A section the registry does not know is NOT skipped quietly -- that is precisely how
    # decided_against went unaudited for its whole life.
    for k, v in data.items():
        if k in _NON_SECTION_KEYS or k in SECTION_SCHEMAS:
            continue
        if isinstance(v, list) and v and all(isinstance(r, dict) for r in v):
            problems.append(
                f"[unknown-section] {k!r} holds {len(v)} row(s) but has no schema in "
                f"SECTION_SCHEMAS -- it would be neither audited nor rendered")
    for sec, exp in sections:
        for r in data.get(sec, []):
            ks = set(r.keys())
            missing, extra = exp - ks, ks - exp
            if missing or extra:
                problems.append(
                    f"[schema] {sec}: {r.get('ID', '?')} "
                    f"missing={sorted(missing)} extra={sorted(extra)}")

    allids = [r.get("ID", "") for sec, _ in sections for r in data.get(sec, [])]
    for i, n in Counter(allids).items():
        if i and n > 1:
            problems.append(f"[dup-id] {i!r} appears {n}x across sections")

    sp = os.path.join(d, "STATE.json")
    if os.path.isfile(sp):
        try:
            tt = json.load(open(sp, encoding="utf-8")).get("task_tracker", "")
            m = re.search(r"(\d+)\s*incomplete\s*/\s*(\d+)\s*awaiting-operator"
                          r"\s*/\s*(\d+)\s*completed", tt)
            if m:
                got = tuple(int(x) for x in m.groups())
                want = (len(data.get("incomplete", [])),
                        len(data.get("awaiting_operator", [])),
                        len(data.get("completed", [])))
                if got != want:
                    problems.append(
                        f"[state-drift] STATE.task_tracker says {got} but "
                        f"DATA.json is {want}")
        except Exception:
            pass

    counts = " / ".join(f"{len(data.get(s, []))} {s}" for s, _ in sections)
    total = sum(len(data.get(s, [])) for s, _ in sections)
    if problems:
        for p in problems:
            print("  FAIL", p)
        print(f"RESULT: DEFECTS ({len(problems)}) -- {counts} = {total} total")
        return 1
    print(f"  [OK] schema conforms | no duplicate IDs | "
          f"counts {counts} = {total} total")
    print("RESULT: CLEAN")
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--bootstrap", action="store_true")
    ap.add_argument("--audit", action="store_true")
    ap.add_argument("--render", action="store_true")
    ap.add_argument("--check", action="store_true")
    ap.add_argument("--status", action="store_true")
    ap.add_argument("dir", nargs="?", default=".")
    a = ap.parse_args(argv)
    if a.bootstrap:
        return bootstrap(a.dir)
    if a.render:
        return render(a.dir)
    if a.check:
        return check(a.dir)
    if a.status:
        return status(a.dir)
    if a.audit:
        return audit(a.dir)
    ap.error("pass --bootstrap <dir> | --render <dir> | --check <dir> | --status <dir> | --audit <dir>")


if __name__ == "__main__":
    raise SystemExit(main())
