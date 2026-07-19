#!/usr/bin/env python3
"""tasktracker_sync — keep TASK_TRACKER.xlsx and TASK_TRACKER.md from drifting.

The canonical task list is TASK_TRACKER.xlsx; TASK_TRACKER.md mirrors it. With
the original generator absent from the pack, the two are hand-synced — and that
is exactly how a row landed in the xlsx but not the md (a GCW-3 drift caught only
by an ad-hoc diff). This tool makes that drift a one-command gate.

  python3 tools/tasktracker_sync.py --check <dir>   # exit 1 if ID sets differ
  python3 tools/tasktracker_sync.py --regen <dir>   # rewrite md tables from xlsx (sole source)
  python3 tools/tasktracker_sync.py --check          # cwd

It compares the ID column of the xlsx Incomplete+Completed sheets against the
IDs in the md tables. Reports rows present in one but not the other.

Requires openpyxl (present in the sandbox + the service venv); it is a
dev/close-time gate, not a stash-runtime chain CLI, so the non-stdlib dep is
acceptable. Fails loudly with an actionable message if openpyxl is missing.
"""
import argparse
import os
import re
import sys


# v3.66.754a: ONE registry, three consumers. This file carried the section list THREE
# times -- xlsx_ids read 2 sheets, md_ids recognised 2 headings, and check() compared 2
# buckets -- while the renderer emitted FOUR sections into FIVE sheets. So the gate that
# build_session_pack.py:114 blocks the pack build on was comparing two DIFFERENTLY-WRONG
# numbers: md swallowed the 26 awaiting rows into Incomplete and the 16 decided rows into
# Completed (measured on the shipped tracker: md=38/374 vs xlsx=12/359). A gate whose two
# sides are both wrong is not a gate -- it can report DRIFT on a tracker that is in sync,
# and pass a real drift that happens to cancel out. Derive from one place, or it drifts again.
SECTIONS = [
    # (canonical name, xlsx sheet, md heading prefix -- lowercase)
    ("Incomplete",        "Incomplete",        "## incomplete"),
    ("Awaiting operator", "Awaiting operator", "## awaiting operator"),
    ("Completed",         "Completed",         "## completed"),
    ("Decided against",   "Decided against",   "## decided against"),
]
_MD_HEADINGS = [(h, name) for name, _sh, h in SECTIONS]


def xlsx_ids(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("tasktracker_sync: openpyxl required (run in the venv/sandbox): "
                 "pip install openpyxl --break-system-packages")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    for _name, sheet, _h in SECTIONS:
        if sheet not in wb.sheetnames:
            continue
        ids = []
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row and row[0] not in (None, ""):
                ids.append(str(row[0]).strip())
        out[_name] = ids
    return out


def md_ids(path):
    out = {name: [] for name, _sh, _h in SECTIONS}
    sheet = None
    with open(path, encoding="utf-8") as fh:
        for ln in fh:
            low = ln.lower()
            if low.startswith("## "):
                # THE FIX: reset on EVERY h2. The old parser only recognised two headings
                # and, on any other, LEFT `sheet` AS-IS -- so it kept appending rows from
                # the next section into the previous bucket. An unrecognised section must
                # make the parser stop, not silently inherit.
                sheet = next((nm for h, nm in _MD_HEADINGS if low.startswith(h)), None)
                continue
            if sheet and ln.startswith("| ") and not re.match(r"\|\s*-+", ln):
                first = ln.split("|")[1].strip()
                # v3.66.754a: the header row used to be skipped with
                # `not ln.startswith("| ID")` -- a PREFIX match. So "| IDEA-HARDEN |"
                # was mistaken for the header and SILENTLY DROPPED. It stayed invisible
                # for as long as it did because BOTH sides of the drift comparison were
                # wrong and the errors cancelled; fixing the denominator surfaced it
                # immediately (md=358 vs xlsx=359). Match the header CELL, not a prefix.
                if first == "ID":
                    continue
                if first:
                    out[sheet].append(first)
    return out


def check(d):
    xp = os.path.join(d, "TASK_TRACKER.xlsx")
    mp = os.path.join(d, "TASK_TRACKER.md")
    if not (os.path.isfile(xp) and os.path.isfile(mp)):
        sys.exit(f"tasktracker_sync: need both files in {d}")
    xi, mi = xlsx_ids(xp), md_ids(mp)
    drift = False
    for sheet, _sh, _h in SECTIONS:
        xs, ms = set(xi.get(sheet, [])), set(mi.get(sheet, []))
        only_x, only_m = sorted(xs - ms), sorted(ms - xs)
        tag = "OK" if not (only_x or only_m) else "DRIFT"
        print(f"  [{tag}] {sheet}: xlsx={len(xs)} md={len(ms)}")
        if only_x:
            print(f"        in xlsx not md: {only_x}")
        if only_m:
            print(f"        in md not xlsx: {only_m}")
        drift = drift or bool(only_x or only_m)
    print("RESULT:", "DRIFT" if drift else "IN-SYNC")
    return 1 if drift else 0


def _rows(path, sheet):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        return [], []
    rows = list(wb[sheet].iter_rows(min_row=1, values_only=True))
    header = [str(c) if c is not None else "" for c in rows[0]] if rows else []
    data = [[("" if c is None else str(c)) for c in r]
            for r in rows[1:] if r and r[0] not in (None, "")]
    return header, data


def _md_table(header, data):
    def esc(s):
        return s.replace("|", "\\|").replace("\n", " ")
    out = ["| " + " | ".join(esc(h) for h in header) + " |",
           "| " + " | ".join("---" for _ in header) + " |"]
    for row in data:
        row = (row + [""] * len(header))[:len(header)]
        out.append("| " + " | ".join(esc(c) for c in row) + " |")
    return "\n".join(out)


def regen(d):
    """Rewrite TASK_TRACKER.md's two tables from the xlsx (sole source),
    preserving everything before the '## Incomplete' heading."""
    xp = os.path.join(d, "TASK_TRACKER.xlsx")
    mp = os.path.join(d, "TASK_TRACKER.md")
    if not os.path.isfile(xp):
        sys.exit(f"tasktracker_sync: missing {xp}")
    ih, idata = _rows(xp, "Incomplete")
    ch, cdata = _rows(xp, "Completed")
    preamble = ""
    if os.path.isfile(mp):
        txt = open(mp, encoding="utf-8").read()
        idx = txt.lower().find("## incomplete")
        preamble = txt[:idx] if idx != -1 else txt
    if not preamble.strip():
        preamble = "# BulkDownloader — unified task tracker\n\n"
    body = (preamble.rstrip() + "\n\n"
            + "## Incomplete (running)\n\n" + _md_table(ih, idata) + "\n\n"
            + "## Completed (running)\n\n" + _md_table(ch, cdata) + "\n")
    open(mp, "w", encoding="utf-8").write(body)
    print(f"  regenerated {mp}: {len(idata)} incomplete, {len(cdata)} completed rows")
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true")
    ap.add_argument("--regen", action="store_true")
    ap.add_argument("dir", nargs="?", default=".")
    a = ap.parse_args(argv)
    if a.regen:
        return regen(a.dir)
    if a.check:
        return check(a.dir)
    ap.error("pass --check <dir> or --regen <dir>")


if __name__ == "__main__":
    raise SystemExit(main())
